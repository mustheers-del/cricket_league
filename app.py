"""
Cricket League Management System - Flask Application
Handles league creation, fixture generation, match results, and points table
"""

from flask import Flask, render_template, request, jsonify, send_file
from database import (
    init_db, create_league, get_league, get_teams, get_fixtures,
    add_match_result, get_points_table, reset_tournament, get_match_by_id,
    update_match_result, delete_match_result
)
import sqlite3
import os
from datetime import datetime
from io import BytesIO
import json

# Initialize Flask app
app = Flask(__name__)
app.config['DATABASE'] = 'cricket_league.db'

# Initialize database on startup
init_db()

# ==================== HELPER FUNCTIONS ====================

def get_db_connection():
    """Create database connection"""
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    return conn

def calculate_nrr(runs_for, overs_faced, runs_against, overs_bowled):
    """Calculate Net Run Rate"""
    if overs_faced == 0 or overs_bowled == 0:
        return 0.0
    
    run_rate_for = runs_for / overs_faced if overs_faced > 0 else 0
    run_rate_against = runs_against / overs_bowled if overs_bowled > 0 else 0
    nrr = run_rate_for - run_rate_against
    return round(nrr, 2)

# ==================== ROUTES ====================

@app.route('/')
def index():
    """Home page"""
    league = get_league()
    if league:
        teams = get_teams()
        points_table = get_points_table()
        return render_template('index.html', league=league, teams=teams, points_table=points_table)
    return render_template('league_setup.html')

@app.route('/api/create-league', methods=['POST'])
def api_create_league():
    """Create a new league"""
    try:
        data = request.json
        league_name = data.get('league_name', '').strip()
        teams = [t.strip() for t in data.get('teams', []) if t.strip()]
        
        # Validation
        if not league_name:
            return jsonify({'success': False, 'error': 'League name is required'}), 400
        
        if len(teams) != 5:
            return jsonify({'success': False, 'error': 'Exactly 5 teams are required'}), 400
        
        if len(set(teams)) != 5:
            return jsonify({'success': False, 'error': 'Team names must be unique'}), 400
        
        # Create league in database
        create_league(league_name, teams)
        
        return jsonify({'success': True, 'message': 'League created successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/league-info')
def api_league_info():
    """Get current league information"""
    league = get_league()
    teams = get_teams()
    
    if not league:
        return jsonify({'has_league': False})
    
    return jsonify({
        'has_league': True,
        'league_name': league['name'],
        'teams': [{'id': t['id'], 'name': t['name']} for t in teams]
    })

@app.route('/fixtures')
def fixtures():
    """Display all fixtures"""
    league = get_league()
    if not league:
        return render_template('league_setup.html')
    
    fixtures = get_fixtures()
    return render_template('fixtures.html', league=league, fixtures=fixtures)

@app.route('/api/fixtures')
def api_fixtures():
    """Get all fixtures as JSON"""
    fixtures = get_fixtures()
    return jsonify(fixtures)

@app.route('/enter-result', methods=['GET', 'POST'])
def enter_result():
    """Enter match result"""

    print("ENTER RESULT ROUTE RUNNING")

    league = get_league()

    if not league:
        return render_template('league_setup.html')


    teams = get_teams()

    fixtures = get_fixtures()


    print("NUMBER OF FIXTURES:", len(fixtures))

    print("ALL FIXTURES:")
    for f in fixtures:
        print(f)


    pending_fixtures = []

    for fixture in fixtures:
        if fixture['result_id'] is None:
            pending_fixtures.append(fixture)


    print("PENDING COUNT:", len(pending_fixtures))


    return render_template(
    'enter_result.html',
    league=league,
    teams=teams,
    pending_fixtures=pending_fixtures
)

@app.route('/api/add-result', methods=['POST'])
def api_add_result():
    """Add match result"""
    try:
        data = request.json
        
        # Validation
        fixture_id = data.get('fixture_id')
        team1_runs = data.get('team1_runs')
        team1_wickets = data.get('team1_wickets')
        team1_overs = data.get('team1_overs')
        team2_runs = data.get('team2_runs')
        team2_wickets = data.get('team2_wickets')
        team2_overs = data.get('team2_overs')
        
        # Type conversion and validation
        try:
            fixture_id = int(fixture_id)
            team1_runs = int(team1_runs)
            team1_wickets = int(team1_wickets)
            team1_overs = float(team1_overs)
            team2_runs = int(team2_runs)
            team2_wickets = int(team2_wickets)
            team2_overs = float(team2_overs)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Invalid input values'}), 400
        
        # Validation checks
        if team1_runs < 0 or team2_runs < 0:
            return jsonify({'success': False, 'error': 'Runs cannot be negative'}), 400
        
        if not (0 <= team1_wickets <= 10) or not (0 <= team2_wickets <= 10):
            return jsonify({'success': False, 'error': 'Wickets must be between 0 and 10'}), 400
        
        if team1_overs < 0 or team2_overs < 0:
            return jsonify({'success': False, 'error': 'Overs cannot be negative'}), 400
        
        # Add result to database
        add_match_result(fixture_id, team1_runs, team1_wickets, team1_overs,
                        team2_runs, team2_wickets, team2_overs)
        
        return jsonify({'success': True, 'message': 'Match result added successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/points-table')
def points_table():
    """Display points table"""
    league = get_league()
    if not league:
        return render_template('league_setup.html')
    
    points_table = get_points_table()
    return render_template('points_table.html', league=league, points_table=points_table)

@app.route('/api/points-table')
def api_points_table():
    """Get points table as JSON"""
    points_table = get_points_table()
    return jsonify(points_table)

@app.route('/api/reset-tournament', methods=['POST'])
def api_reset_tournament():
    """Reset entire tournament"""
    try:
        reset_tournament()
        return jsonify({'success': True, 'message': 'Tournament reset successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export-excel')
def export_excel():
    """Export points table to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        league = get_league()
        points_table = get_points_table()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Points Table"
        
        # Add title
        ws['A1'] = f"{league['name']} - Points Table"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:K1')
        
        # Add headers
        headers = ['Rank', 'Team', 'Matches', 'Wins', 'Losses', 'Points', 
                   'Runs For', 'Runs Against', 'Overs Faced', 'Overs Bowled', 'NRR']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for row_idx, team in enumerate(points_table, 4):
            ws.cell(row=row_idx, column=1).value = team['rank']
            ws.cell(row=row_idx, column=2).value = team['team_name']
            ws.cell(row=row_idx, column=3).value = team['matches']
            ws.cell(row=row_idx, column=4).value = team['wins']
            ws.cell(row=row_idx, column=5).value = team['losses']
            ws.cell(row=row_idx, column=6).value = team['points']
            ws.cell(row=row_idx, column=7).value = team['runs_for']
            ws.cell(row=row_idx, column=8).value = team['runs_against']
            ws.cell(row=row_idx, column=9).value = team['overs_faced']
            ws.cell(row=row_idx, column=10).value = team['overs_bowled']
            ws.cell(row=row_idx, column=11).value = team['nrr']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 12
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'cricket_league_points_table.xlsx'
        )
    
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-pdf')
def export_pdf():
    """Export points table to PDF"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        
        league = get_league()
        points_table = get_points_table()
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        title = Paragraph(f"{league['name']} - Points Table", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Create table data
        data = [['Rank', 'Team', 'M', 'W', 'L', 'Pts', 'RF', 'RA', 'OF', 'OB', 'NRR']]
        
        for team in points_table:
            data.append([
                str(team['rank']),
                team['team_name'],
                str(team['matches']),
                str(team['wins']),
                str(team['losses']),
                str(team['points']),
                str(team['runs_for']),
                str(team['runs_against']),
                str(team['overs_faced']),
                str(team['overs_bowled']),
                str(team['nrr'])
            ])
        
        # Create table
        table = Table(data, colWidths=[0.5*inch]*11)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')])
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'cricket_league_points_table.pdf'
        )
    
    except ImportError:
        return jsonify({'error': 'reportlab not installed'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/edit-result/<int:match_id>')
def edit_result(match_id):
    """Edit match result page"""
    league = get_league()
    if not league:
        return render_template('league_setup.html')
    
    match = get_match_by_id(match_id)
    if not match:
        return "Match not found", 404
    
    return render_template('edit_result.html', league=league, match=match)

@app.route('/api/match/<int:match_id>')
def api_get_match(match_id):
    """Get match details as JSON"""
    match = get_match_by_id(match_id)
    if not match:
        return jsonify({'error': 'Match not found'}), 404
    return jsonify(match)

@app.route('/api/update-result/<int:match_id>', methods=['POST'])
def api_update_result(match_id):
    """Update match result"""
    try:
        data = request.json
        
        team1_runs = int(data.get('team1_runs', 0))
        team1_wickets = int(data.get('team1_wickets', 0))
        team1_overs = float(data.get('team1_overs', 0))
        team2_runs = int(data.get('team2_runs', 0))
        team2_wickets = int(data.get('team2_wickets', 0))
        team2_overs = float(data.get('team2_overs', 0))
        
        update_match_result(match_id, team1_runs, team1_wickets, team1_overs,
                          team2_runs, team2_wickets, team2_overs)
        
        return jsonify({'success': True, 'message': 'Match updated successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/delete-result/<int:match_id>', methods=['POST'])
def api_delete_result(match_id):
    """Delete match result"""
    try:
        delete_match_result(match_id)
        return jsonify({'success': True, 'message': 'Match result deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(error):
    """Handle 500 errors"""
    return render_template('500.html'), 500

# ==================== MAIN ====================

if __name__ == '__main__':
    app.run(
    host="0.0.0.0",
    port=5000,
    debug=True
)